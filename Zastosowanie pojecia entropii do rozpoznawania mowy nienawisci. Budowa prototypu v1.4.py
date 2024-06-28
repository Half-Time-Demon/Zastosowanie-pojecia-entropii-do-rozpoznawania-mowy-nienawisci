import math
import os
from tkinter import Tk    
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side

#clear text off neutral words and signs + 
def clear_text(text_data): 
    try:   
        for key in text_data.keys():
            if ('text' in key and any(map(str.isdigit, key))):
                text_data[key]['text_clear'] = text_data[key]['text'].lower()
                words = text_data[key]['text_clear'].split()
                clear_text = ""
                text_data[key]['text_word_count'] = 0
                text_data[key]['clear_text_word_count'] = 0

                for word in words:
                    text_data[key]['text_word_count'] += 1
                    clear_word = word.strip('.,?!()[]{}";:\'=+-_/\\')
                    if clear_word in text_data['neutral_words_dict']:
                        continue
                    clear_text += clear_word +" "
                    text_data[key]['clear_text_word_count'] += 1

                text_data[key]['text_clear'] = clear_text
    except Exception as e:
        clear()
        print(f"Error in clear_text: {e}")
        print("")
        input("press any key to continue >>>")

#counts from hate speach dictionary words in text +
def check_hate_speach(text_data):
    try:
        for key in text_data.keys():
            if ('text' in key and any(map(str.isdigit, key))):
                if('text_clear' not in text_data[key].keys()):
                    clear_text(text_data)
                text_data[key]['hate_speach_counts'] =  [text_data[key]['text_clear'].count(word) for word in text_data['hate_speach_dict']]
    except Exception as e:
        clear()
        print(f"Error in check_hate_speach: {e}")
        print("")
        input("press any key to continue >>>")

# colculate entropy for every word whose fanded in a text +   
def calculate_entropy(text_data):
    try:
        for key in text_data.keys():
            if ('text' in key and any(map(str.isdigit, key))):
                hate_speach_entropy = []
                if ('hate_speach_counts' not in text_data.keys()):
                    check_hate_speach(text_data)

                for num in text_data[key]['hate_speach_counts']:
                    if num == 0:
                        hate_speach_entropy.append(0)
                        continue
                    
                    hate_speach_entropy.append(-(num/text_data[key]['clear_text_word_count']*math.log(num/text_data[key]['clear_text_word_count'],2)))
                for S in hate_speach_entropy:
                    if '-' in str(S):
                        hate_speach_entropy[hate_speach_entropy.index(S)] = 1
                text_data[key]['hate_speach_entropy'] = hate_speach_entropy  
    except Exception as e:
        clear()
        print(f"Error in calculate_entropy: {e}")
        print("")
        input("press any key to continue >>>")  

# colculate sum of all entropy and check level of entropy in text +
def sum_of_entropy(text_data,entropy_level=0.2):
    try:
        for key in text_data.keys():
            if ('text' in key and any(map(str.isdigit, key))):
                if ('hate_speach_entropy' not in text_data[key].keys()):
                    calculate_entropy(text_data)
                text_data[key]['sum_of_entropy'] = sum(text_data[key]['hate_speach_entropy'])
                if text_data[key]['sum_of_entropy'] >= entropy_level:
                    text_data[key]['hate_speach'] = True
                else:
                    text_data[key]['hate_speach'] = False
    except Exception as e:
        clear()
        print(f"Error in sum_of_entropy: {e}")
        print("")
        input("press any key to continue >>>")

#check if directoris exist in program folder if not create him +
def check_make_dirs():
    try:
        if not os.path.exists('Hate_speach_dict'):
             os.mkdir('Hate_speach_dict')
        if not os.path.exists('Neutral_words_dict'):
             os.mkdir('Neutral_words_dict')
        if not os.path.exists('Text_files'):
             os.mkdir('Text_files')
        if not os.path.exists('Result'):
             os.mkdir('Result')  
    except Exception as e:
        clear()
        print(f"Error in check_make_dirs: {e}") 
        print("")
        input("press any key to continue >>>")

#save results in excel file +
def save_result(text_data,fn="result"):
    try:
        fp = os.path.join(os.getcwd(), "Result", f"{fn}.xlsx")
        if not os.path.isfile(fp):
            wb = Workbook()
        else:
            wb = load_workbook(fp)
            ws = wb.active
            wb.remove(ws)
            ws = wb.create_sheet(title='Data')
        
        #create datasheet, create info line and fill it
        ws = wb.active
        ws.title = 'Data'
        ws.append([''])
        info_line=[]
        for col_name in text_data[list(text_data.keys())[0]].keys():
            if col_name == "hate_speach_entropy" or col_name == "hate_speach_counts":
                for i in text_data['hate_speach_dict']:
                    info_line.append(i)
            else:
                info_line.append(col_name)
        ws.append(info_line)
        
        #decoration
        header_fill_1 = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        header_fill_2 = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
        header_fill_3 = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        header_fill_4 = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        header_fill_5 = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        header_font_1 = Font(color="CF7200", bold=True)
        header_font_2 = Font(color="FFFFF1", bold=True)
        header_font_3 = Font(color="9C0031", bold=True)
        header_font_4 = Font(color="6F7B00", bold=True)
        header_font_5 = Font(color="CF7200", bold=True)
        
        answer_color_red_font = Font(color="E24539", bold=True)
        answer_color_green_font = Font(color="6D9542", bold=True)
        entropy_color_orange_font = Font(color="F87D00", bold=True)
        
        info_border_left = Border(
            left=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        info_border_right = Border(
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        info_border_middle = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        max_length = 0
        main_col_num = len(text_data[list(text_data.keys())[0]].keys())
        sec_col_num = len(text_data['hate_speach_dict'])*2
        
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        for column_number in range(1, (main_col_num+sec_col_num)-1):
            for row in ws.iter_rows(min_row=2, max_row=2, min_col=column_number, max_col=column_number):
                for cell in row:
                    if column_number <= 4:
                        cell.fill = header_fill_1
                        cell.font = header_font_1
                    elif column_number <= 4 + (sec_col_num/2):
                        cell.fill = header_fill_2
                        cell.font = header_font_2
                    elif column_number <= 4 + sec_col_num:
                        cell.fill = header_fill_3
                        cell.font = header_font_3
                    elif column_number == 5 + sec_col_num:
                        cell.fill = header_fill_4
                        cell.font = header_font_4
                    else: 
                        cell.fill = header_fill_5
                        cell.font = header_font_5
                        
                    if column_number == 1:
                        cell.border = info_border_left
                    elif column_number < (main_col_num + sec_col_num -2):
                        cell.border = info_border_middle
                    else:
                        cell.border = info_border_right
        
        #fill information in file
        for row in range(1,text_data['text_count']+1):
            for col in range(1,len(text_data[list(text_data.keys())[0]].keys())+(len(text_data['hate_speach_dict'])*2)-1):
                if col >= 5 and col < (5+len(text_data['hate_speach_dict'])):
                    ws.cell(row=row+2,column=col,value=text_data['text'+str(row)]['hate_speach_counts'][col-5])
                elif col >= (5+len(text_data['hate_speach_dict'])) and col < (5+(len(text_data['hate_speach_dict'])*2)):
                    ws.cell(row=row+2,column=col,value=text_data['text'+str(row)]['hate_speach_entropy'][col-5-len(text_data['hate_speach_dict'])])
                elif col < 5:  
                    ws.cell(row=row+2,column=col,value=text_data['text'+str(row)][list(text_data['text'+str(row)].keys())[col-1]])
                else:
                    ws.cell(row=row+2,column=col,value=text_data['text'+str(row)][list(text_data['text'+str(row)].keys())[col+1-len(text_data['hate_speach_dict'])*2]])
          
        #decoration              
        for column_number in range(1, (main_col_num+sec_col_num)-1):
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=5, max_col=column_number):
                for cell in row:
                    if cell.value > 0:
                        cell.font = entropy_color_orange_font
                    
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=ws.max_column, max_col=ws.max_column):
            for cell in row:
                if cell.value == True:
                    cell.font = answer_color_green_font
                else:
                    cell.font = answer_color_red_font

        wb.save(fp)
        wb.close()
        
        print("succesfuly saved!")
        input("press enter to continue >>>")
    except IndexError as e:
        clear()
        print(f"Error in save_result: {e}")
        print("without \"sum of entropy\" you can't save result!!!")
        print("")
        input("press any key to continue >>>")
    except Exception as e:
        clear()
        print(f"Error in save_result: {e}")
        print("")
        input("press any key to continue >>>")

#check or file by path is empty +
def file_is_empty(path):
    try:
        return os.stat(path).st_size==0
    except Exception as e:
        clear()
        print(f"Error in file_is_empty: {e}")
        print("")
        input("press any key to continue >>>")

#this function need for reset those two masives text_data +
def reset_data(text_data):
    try:
        text_data={}
        
    except Exception as e:
        clear()
        print(f"Error in reset_data: {e}")
        print("")
        input("press any key to continue >>>")

# load text data with text for manipulations + 
def load_text_file(text_data, filename=""):
    try:
        if filename == "":
            Tk().withdraw()
            filename = askopenfilename(initialdir=os.getcwd()+r"\Text_files")
        with open(filename, "r", encoding="utf-8") as file:
            text = file.read()
            text_list = text.split("\n\n")
            i=0
            for text in text_list:
                i+=1
                text_data['text'+str(i)]={'text':text.replace("\n"," ")}
            text_data['text_count']=i
        return filename
    except Exception as e:
        clear()
        print(f"Error in load_text_file: {e}")
        print("")
        input("press any key to continue >>>")

# load hate speach dictionary +
def load_hate_speach_dict(text_data, filename=""):
    try:
        if filename == "":
            Tk().withdraw()
            filename = askopenfilename(initialdir=os.getcwd()+r"\Hate_speach_dict")
        with open(filename, "r", encoding="utf-8") as file:
            text = file.read()
            text_data['hate_speach_dict'] = text.split("\n")
        return filename
    except Exception as e:
        clear()
        print(f"Error in load_hate_speach_dict: {e}")
        print("")
        input("press any key to continue >>>")

# load neutral words dictionary +     
def load_neutral_words_dict(text_data, filename=""):
    try:
        if filename == "":
            Tk().withdraw()
            filename = askopenfilename(initialdir=os.getcwd()+r"\Neutral_words_dict")
        with open(filename, "r", encoding="utf-8") as file:
            text = file.read()
            text_data['neutral_words_dict'] = text.split("\n")
        return filename
    except Exception as e:
        clear()
        print(f"Error in load_neutral_words_dict: {e}")
        print("")
        input("press any key to continue >>>")

#create settings.ini in program folder and fill it +
def create_settings():
    try:
        with open('settings.ini', "a", encoding="utf-8") as file:
        
            clear()
            input("press enter to load txt file >>>")
            file.write("text_file="+load_text_file(text_data)+"\n")

            clear()
            input("press enter to load the hate speach dictionary >>>")
            file.write("hate_speach_dict="+load_hate_speach_dict(text_data)+"\n")

            clear()
            input("press enter to load the neutral words dictionary >>>")
            file.write("neutral_words_dict="+load_neutral_words_dict(text_data))

    except Exception as e:
        clear()
        print(f"Error in create_settings: {e}")
        print("")
        input("press any key to continue >>>")

#function that load paths from settings.ini +
def load_settings(settings_paths, text_data):
    try:
        if not file_is_empty('settings.ini'):
            with open('settings.ini', "r", encoding="utf-8") as file:
                settings = file.read().split("\n")
                for s in settings:
                    s=s.split("=")
                    settings_paths[s[0]]=s[1]

            load_text_file(text_data,settings_paths['text_file'])
            load_hate_speach_dict(text_data,settings_paths['hate_speach_dict'])
            load_neutral_words_dict(text_data,settings_paths['neutral_words_dict'])
            input("setting is loaded >>>")
        else:
            print("settings file is empty!!!")
            create_settings()
    except Exception as e:
        clear()
        print(f"Error in load_settings: {e}")
        print("")
        input("press any key to continue >>>")
 
#function for save and rewrite settings.ini after changes +
def save_settings(settings_paths):
    try:
        with open('settings.ini', "a", encoding="utf-8") as file:
            for k in settings_paths:
                file.write(k+'='+settings_paths[k])

        reset_data(text_data)
        load_settings(settings_paths, text_data)
    except Exception as e:
        clear()
        print(f"Error in save_settings: {e}")
        print("")
        input("press any key to continue >>>")

# this function load other functions. Those need for correctli working entire program +       
def start():
    try:
        check_make_dirs()

        if not os.path.exists('settings.ini'):
            create_settings()
        else:
            load_settings(settings_paths, text_data)
    except Exception as e:
        clear()
        print(f"Error in start: {e}")
        print("")
        input("press any key to continue >>>")
    

#program starts from this position    
settings_paths={}
text_data={}
clear = lambda: os.system('cls')

start()
    
#menu +
while True:
    clear()
    print("1.Check text on hate speach")
    print("2.Check entropy for hate speach")
    print("3.Sum of entropy")
    print("4.Text data")
    print("5.settings")
    print("6.Save result in excel")
    print("0.Exit")
    term = input(">>>")
    
    match term:
        case "1":
            clear()
            check_hate_speach(text_data)
            for key in text_data.keys():
                if ('text' in key and any(map(str.isdigit, key))):
                    print("hate speach counters " + key + " : " + str(text_data[key]['hate_speach_counts']))
            input("press enter to continue >>>")
        case "2":
            clear()
            calculate_entropy(text_data)
            for key in text_data.keys():
                if ('text' in key and any(map(str.isdigit, key))):
                    print("hate speach entropii " + key + " : " + str(text_data[key]['hate_speach_entropy']))
            input("press enter to continue >>>")
        case "3":
            clear()
            sum_of_entropy(text_data)
            for key in text_data.keys():
                if ('text' in key and any(map(str.isdigit, key))): 
                    print("suma entropii " + key + " : " + str(text_data[key]['sum_of_entropy']))
            input("press enter to continue >>>")
        case "4":
            clear()
            print(text_data)
            input("press enter to continue >>>")
        case "5":
            #sub menu for settings
            while True:
                clear()
                print("1.Change/Choose hate speach dictionary")
                print("2.Change/Choose neutral words dictionary")
                print("3.Change/Choose txt file")
                print("0.back")
                term = input(">>>")
    
                match term:
                    case "1":
                        clear()
                        input("press enter to load the hate speach dictionary >>>")
                        settings_paths['hate_speach_dict'] = load_hate_speach_dict(text_data)
                    case "2":
                        clear()
                        input("press enter to load txt neutral words dictionary >>>")
                        settings_paths['neutral_words_dict'] = load_neutral_words_dict(text_data)
                    case "3":
                        clear()
                        input("press enter to load txt file >>>")
                        settings_paths['text_file'] = load_text_file(text_data)
                    case "0":
                        save_settings(settings_paths)
                        break
                    case _:
                        clear()
                        print("wrong number of the menu")
                        input("press enter to continue >>>")
        case "6":
            clear()
            fn=input("input file name of file (dont write .xlsx or some thing else) >>>")    
            if fn:
                save_result(text_data,fn)
            else:
                save_result(text_data)
        case "0":
            clear()
            quit = input("press Y-to exit N-go back >>>")
            if quit.upper() == "Y":
                clear()
                break
            elif quit.upper() == "N":
                pass
            else:
                print("wrong input")
        case _:
            clear()
            print("wrong number of the menu")
            input("press enter to continue >>>")
       